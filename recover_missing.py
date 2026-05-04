#!/usr/bin/env python3
"""Recovery script to capture missing actual outputs for specific rows."""

from playwright.sync_api import sync_playwright
import re
import time
from openpyxl import load_workbook
from pathlib import Path

def test_inputs():
    problem_rows = [2, 36, 37, 38, 42, 43, 46, 47, 51]
    
    # Load Excel to get input values
    excel_path = Path(r'C:/Users/User/Desktop/test_automation/test_automation/Assignment 1 - Test cases.xlsx')
    wb = load_workbook(excel_path, keep_vba=False, rich_text=False)
    ws = wb[' Test cases']
    
    inputs_to_test = []
    for row_num in problem_rows:
        input_text = ws.cell(row_num, 3).value
        if input_text:
            inputs_to_test.append((row_num, input_text))
    
    wb.close()
    
    # Start browser and test each input
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        page = browser.new_page()
        page.goto("https://www.pixelssuite.com/chat-translator", wait_until="domcontentloaded", timeout=15000)
        
        # Wait for page to load
        page.wait_for_timeout(2000)
        
        results = {}
        
        for row_num, input_text in inputs_to_test:
            print(f"\nTesting Row {row_num}: {input_text[:50]}")
            
            try:
                # Find and clear input
                input_locator = page.locator("textarea").first
                input_locator.clear()
                input_locator.click()
                
                # Type input with delay
                input_locator.type(input_text, delay=20)
                
                # Find and click Transliterate button
                button = page.locator("button").filter(has_text=re.compile(r"Transliterate", re.I)).last
                
                for attempt in range(3):
                    try:
                        button.click(timeout=1000)
                        break
                    except:
                        page.wait_for_timeout(300)
                
                # Wait for output
                page.wait_for_timeout(3000)
                
                # Try to extract output from output textarea
                output_locator = page.locator("textarea").nth(1)
                actual_output = output_locator.input_value().strip() if output_locator else ""
                
                # If empty, try to find it in the card div
                if not actual_output:
                    card_text = page.locator("div[role='main']").text_content()
                    if card_text:
                        actual_output = card_text.strip()
                
                results[row_num] = actual_output if actual_output else "[NO OUTPUT CAPTURED]"
                print(f"  -> Captured: {results[row_num][:50]}")
                
            except Exception as e:
                results[row_num] = f"[ERROR: {str(e)[:30]}]"
                print(f"  -> Error: {e}")
        
        browser.close()
    
    # Update Excel with results
    wb = load_workbook(excel_path, keep_vba=False, rich_text=False)
    ws = wb[' Test cases']
    
    for row_num, actual in results.items():
        ws.cell(row_num, 5).value = actual  # Column E = Actual output
    
    wb.save(excel_path)
    wb.close()
    
    print("\n[DONE] Updated Excel with recovered outputs")

if __name__ == "__main__":
    test_inputs()
